from django.core.management.base import BaseCommand
import os, io
from openpyxl import Workbook
from django.urls import exceptions
from dotenv import load_dotenv
import telebot
from telebot import types
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict
from django.utils.timezone import now, localtime
from shop.models import ProductHistory, Order
from user.models import User

load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN')
bot = telebot.TeleBot(BOT_TOKEN)

user_data = {}
user_message_ids = {}


@bot.message_handler(commands=['start'])
def handle_start(message):
    user_id = str(message.from_user.id)
    user_exists = User.objects.filter(user_telegram_id=user_id).first()

    if not user_exists:
        bot.send_message(user_id, "Ismingizni kiriting: ")
        bot.register_next_step_handler(message, get_user_info)
    elif user_exists.is_superuser:
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
        buttons = [
            types.InlineKeyboardButton("🔥 Bosh menyu", web_app=web_app_info),
            types.InlineKeyboardButton("📆 Shu oygi buyurtmalar", callback_data='orders_this_month'),
            types.InlineKeyboardButton("📋 Barcha buyurtmalar", callback_data='orders_all'),
            types.InlineKeyboardButton("📆 Shu oygi mahsulotlar", callback_data='products_this_month'),
            types.InlineKeyboardButton("📦 Barcha mahsulotlar", callback_data='products_all'),
        ]
        keyboard.add(*buttons)
        bot.send_message(user_id, "Assalomu alaykum! Bo'limlardan birini tanlang.", reply_markup=keyboard)
    else:
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
        buttons = [
            types.InlineKeyboardButton("🔥 Mahsulotlar", web_app=web_app_info),
            types.InlineKeyboardButton("👤 Siz haqingizda", callback_data='user'),
            types.InlineKeyboardButton(" 📞 Bog'lanish", callback_data='support'),
        ]
        keyboard.add(*buttons)
        bot.send_message(user_id, "Assalomu alaykum! Tanlang:", reply_markup=keyboard)


@bot.callback_query_handler(func=lambda call: True)
def handle_inline_buttons(call):
    user_id = call.from_user.id

    try:
        bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    except Exception as e:
        print("Xabar o‘chirishda xatolik:", e)

    if call.data == 'support':
        menu_keyboard = types.InlineKeyboardMarkup(row_width=1)
        menu_keyboard.add(types.InlineKeyboardButton('⬅️ orqaga', callback_data="orqaga"))
        message_text = (
            "📞 <b>Aloqa uchun ma'lumotlar:</b>\n"
            "━━━━━━━━━━━━━━━━━━\n"
            "📱 <b>Telefon:</b> +998999774977\n"
            "📱 <b>Telefon:</b> +998770677717\n"
            "💬 <b>Telegram:</b> <a href='https://t.me/Umarhon999774977'>@Umarhon999774977</a>\n\n"
            "📍 <b>Manzil:</b>\n"
            "<a href='https://www.google.com/maps/place//@41.00921,71.667123,15z'>📌 Xaritada ko‘rish (Google Maps)</a>\n"
            "━━━━━━━━━━━━━━━━━━"
        )

        bot.send_message(
            user_id,
            message_text,
            reply_markup=menu_keyboard,
            parse_mode="HTML"
        )

    elif call.data == 'orqaga':
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
        buttons = [
            types.InlineKeyboardButton("🔥 Mahsulotlar", web_app=web_app_info),
            types.InlineKeyboardButton("👤 Siz haqingizda", callback_data='user'),
            types.InlineKeyboardButton(" 📞 Bog'lanish", callback_data='support'),
        ]
        keyboard.add(*buttons)
        bot.send_message(user_id, "Assalomu alaykum. Tanlang:", reply_markup=keyboard)
    elif call.data == 'user':
        user = User.objects.filter(user_telegram_id=user_id).first()
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        buttons = [
            types.InlineKeyboardButton("🔄 Malumotlarni yangilash", callback_data="update_info"),
            types.InlineKeyboardButton('⬅️ Orqaga', callback_data="orqaga"),
        ]
        keyboard.add(*buttons)
        bot.send_message(user_id, f"Isim: {user.first_name} \nTelefon raqam: {user.phone_number} ", reply_markup=keyboard)
    elif call.data == 'update_info':
        bot.send_message(user_id, "Ismingizni kiriting: ")
        bot.register_next_step_handler(call.message, get_user_info)
    elif call.data == 'orders_this_month':
        user_id = call.from_user.id
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Shu oygi buyurtmalar"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")

            headers = ['#', 'Buyurtmachi', 'Buyurmachi telefon raqami', 'Buyurtma holati', 'Mahsulot nomi', 'SKU', 'Bo\'lim', 'Soni', 'Xarajat narxi(dona)', 'Sotuv narxi(dona)',
                       'Jami narx', 'Yaratilgan vaqt']
            ws.append(headers)
            for cell in ws["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            column_widths = [5, 17, 27, 20, 20, 20, 15, 15, 15, 15, 15, 22]
            for i, width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = width
            today = now()
            start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            orders = Order.objects.filter(created_at__gte=start_of_month).order_by('created_at')

            total_revenue = 0
            total_cost = 0
            counter = 1

            for order in orders:
                for item in order.order_items.all():
                    product = item.product
                    quantity = item.quantity or 0
                    sell_price = item.price or 0
                    received_price = product.price_received or 0

                    total_price = sell_price * quantity
                    cost_price = received_price * quantity

                    total_revenue += total_price
                    total_cost += cost_price

                    ws.append([
                        counter,
                        order.user.first_name or '',
                        order.user.phone_number or '',
                        order.status,
                        product.name or '',
                        product.sku or '',
                        str(product.category) if hasattr(product, 'category') else '',
                        quantity,
                        received_price,
                        sell_price,
                        total_price,
                        localtime(order.created_at).strftime("%Y-%m-%d %H:%M:%S")
                    ])
                    counter += 1
            ws2 = wb.create_sheet(title="Hisobot")
            ws2.append(["Buyurtma holati", "Buyurtmalar soni", "Xarajat", "Tushum", "Soft foyda"])
            for cell in ws2["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            column_widths = [20, 20, 20, 20, 20]
            for i, width in enumerate(column_widths, start=1):
                col_letter = ws2.cell(row=1, column=i).column_letter
                ws2.column_dimensions[col_letter].width = width
            status_summary = {
                'tasdiqlandi': {'count': 0, 'revenue': 0, 'cost': 0},
                'jarayonda': {'count': 0, 'revenue': 0, 'cost': 0},
                'bekor qilindi': {'count': 0, 'revenue': 0, 'cost': 0},
            }

            for order in orders:
                status = order.status
                for item in order.order_items.all():
                    product = item.product
                    quantity = abs(item.quantity or 0)
                    sell_price = abs(item.price or 0)
                    received_price = abs(product.price_received or 0)

                    total_price = sell_price * quantity
                    cost_price = received_price * quantity

                    status_summary[status]['count'] += 1
                    status_summary[status]['revenue'] += total_price
                    status_summary[status]['cost'] += cost_price
            for status, data in status_summary.items():
                profit = abs(data['revenue'] - data['cost'])
                ws2.append([
                    status.title(),
                    data['count'],
                    data['cost'],
                    data['revenue'],
                    profit
                ])
            total_orders = sum(data['count'] for data in status_summary.values())
            total_revenue = sum(data['revenue'] for data in status_summary.values())
            total_cost = sum(data['cost'] for data in status_summary.values())
            total_profit = abs(total_revenue - total_cost)

            ws2.append([])
            ws2.append(["Jami", total_orders, total_cost, total_revenue, total_profit])
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            bot.send_document(
                user_id,
                excel_file,
                visible_file_name=f"shu_oy_buyurtmalar_{today.strftime('%Y_%m_%d_%H_%M')}.xlsx",
                caption="📦 Shu oygi buyurtmalar va foyda hisobot fayli"
            )
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
            buttons = [
                types.InlineKeyboardButton("🔥 Bosh menyu", web_app=web_app_info),
                types.InlineKeyboardButton("📆 Shu oygi buyurtmalar", callback_data='orders_this_month'),
                types.InlineKeyboardButton("📋 Barcha buyurtmalar", callback_data='orders_all'),
                types.InlineKeyboardButton("📆 Shu oygi mahsulotlar", callback_data='products_this_month'),
                types.InlineKeyboardButton("📦 Barcha mahsulotlar", callback_data='products_all'),
            ]
            keyboard.add(*buttons)
            bot.send_message(user_id, "✅ Hisobot yuborildi. Menyudan birini tanlang:", reply_markup=keyboard)
        except Exception as e:
            print("Xatolik:", e)
            bot.send_message(user_id, "❌ Buyurtmalar hisobotida xatolik yuz berdi.")

    elif call.data == 'orders_all':
        user_id = call.from_user.id
        try:
            wb = Workbook()
            del wb["Sheet"]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")

            orders = Order.objects.all().order_by("created_at")
            monthly_data = defaultdict(list)

            for order in orders:
                month_key = localtime(order.created_at).strftime("%Y-%m")
                monthly_data[month_key].append(order)

            total_summary = {
                "tasdiqlandi": {"count": 0, "revenue": 0, "cost": 0},
                "jarayonda": {"count": 0, "revenue": 0, "cost": 0},
                "bekor qilindi": {"count": 0, "revenue": 0, "cost": 0},
            }

            for month, orders_list in monthly_data.items():
                ws = wb.create_sheet(title=month)
                headers = ['#', 'Buyurtmachi', 'Buyurmachi telefon raqami', 'Buyurtma holati', 'Mahsulot nomi', 'SKU',
                           'Bo\'lim', 'Soni', 'Xarajat narxi(dona)', 'Sotuv narxi(dona)', 'Jami narx', 'Yaratilgan vaqt']
                ws.append(headers)

                for cell in ws["1:1"]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align

                column_widths = [5, 17, 27, 20, 20, 20, 15, 15, 15, 15, 15, 22]
                for i, width in enumerate(column_widths, start=1):
                    col_letter = ws.cell(row=1, column=i).column_letter
                    ws.column_dimensions[col_letter].width = width

                counter = 1
                monthly_summary = {
                    "tasdiqlandi": {"count": 0, "revenue": 0, "cost": 0},
                    "jarayonda": {"count": 0, "revenue": 0, "cost": 0},
                    "bekor qilindi": {"count": 0, "revenue": 0, "cost": 0},
                }

                for order in orders_list:
                    status = order.status
                    for item in order.order_items.all():
                        product = item.product
                        quantity = abs(item.quantity or 0)
                        sell_price = abs(item.price or 0)
                        received_price = abs(product.price_received or 0)

                        total_price = sell_price * quantity
                        cost_price = received_price * quantity

                        monthly_summary[status]['count'] += 1
                        monthly_summary[status]['revenue'] += total_price
                        monthly_summary[status]['cost'] += cost_price

                        total_summary[status]['count'] += 1
                        total_summary[status]['revenue'] += total_price
                        total_summary[status]['cost'] += cost_price

                        ws.append([
                            counter,
                            order.user.first_name or '',
                            order.user.phone_number or '',
                            order.status,
                            product.name or '',
                            product.sku or '',
                            str(product.category) if hasattr(product, 'category') else '',
                            quantity,
                            received_price,
                            sell_price,
                            total_price,
                            localtime(order.created_at).strftime("%Y-%m-%d %H:%M:%S")
                        ])
                        counter += 1

                # Create monthly summary sheet
                ws_sum = wb.create_sheet(title=f"{month}_hisobot")
                ws_sum.append(["Buyurtma holati", "Buyurtmalar soni", "Xarajat", "Tushum", "Soft foyda"])
                for cell in ws_sum["1:1"]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align

                column_widths = [20, 20, 20, 20, 20]
                for i, width in enumerate(column_widths, start=1):
                    col_letter = ws_sum.cell(row=1, column=i).column_letter
                    ws_sum.column_dimensions[col_letter].width = width

                for status, data in monthly_summary.items():
                    profit = abs(data["revenue"] - data["cost"])
                    ws_sum.append([
                        status.title(),
                        data["count"],
                        data["cost"],
                        data["revenue"],
                        profit
                    ])

                total_orders = sum(d["count"] for d in monthly_summary.values())
                total_revenue = sum(d["revenue"] for d in monthly_summary.values())
                total_cost = sum(d["cost"] for d in monthly_summary.values())
                total_profit = abs(total_revenue - total_cost)
                ws_sum.append([])
                ws_sum.append(["Jami", total_orders, total_cost, total_revenue, total_profit])

            # Create total summary sheet
            ws_total = wb.create_sheet(title="Umumiy_hisobot")
            ws_total.append(["Buyurtma holati", "Buyurtmalar soni", "Xarajat", "Tushum", "Soft foyda"])
            for cell in ws_total["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            for i, width in enumerate(column_widths, start=1):
                col_letter = ws_total.cell(row=1, column=i).column_letter
                ws_total.column_dimensions[col_letter].width = width

            for status, data in total_summary.items():
                profit = abs(data["revenue"] - data["cost"])
                ws_total.append([
                    status.title(),
                    data["count"],
                    data["cost"],
                    data["revenue"],
                    profit
                ])

            grand_orders = sum(d["count"] for d in total_summary.values())
            grand_revenue = sum(d["revenue"] for d in total_summary.values())
            grand_cost = sum(d["cost"] for d in total_summary.values())
            grand_profit = abs(grand_revenue - grand_cost)
            ws_total.append([])
            ws_total.append(["Jami", grand_orders, grand_cost, grand_revenue, grand_profit])

            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            today = now()
            bot.send_document(
                user_id,
                excel_file,
                visible_file_name=f"barcha_oy_buyurtmalar_{today.strftime('%Y_%m_%d_%H_%M')}.xlsx",
                caption="📦 Barcha oylar bo‘yicha buyurtmalar va foyda hisobot fayli"
            )

            keyboard = types.InlineKeyboardMarkup(row_width=1)
            web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
            buttons = [
                types.InlineKeyboardButton("🔥 Bosh menyu", web_app=web_app_info),
                types.InlineKeyboardButton("📆 Shu oygi buyurtmalar", callback_data='orders_this_month'),
                types.InlineKeyboardButton("📋 Barcha buyurtmalar", callback_data='orders_all'),
                types.InlineKeyboardButton("📆 Shu oygi mahsulotlar", callback_data='products_this_month'),
                types.InlineKeyboardButton("📦 Barcha mahsulotlar", callback_data='products_all'),
            ]
            keyboard.add(*buttons)
            bot.send_message(user_id, "✅ Barcha oylar bo‘yicha hisobot yuborildi. Menyudan birini tanlang:",
                             reply_markup=keyboard)

        except Exception as e:
            print("Xatolik:", e)
            bot.send_message(user_id, "❌ Hisobot faylini yaratishda xatolik yuz berdi.")

    elif call.data == 'products_this_month':
        user_id = call.from_user.id
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Bu oygi mahsulotlar"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            headers = ['#', 'Maxsulot nomi', 'SKU', "Bo'lim", 'Olingan narx', 'Sotiladigan narx',
                       'Amal', 'Soni', "Yaratilgan vaqt"]
            ws.append(headers)
            for cell in ws["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            column_widths = [5, 20, 20, 20, 15, 15, 15, 15, 25]
            for i, width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = width

            today = now()
            start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            products = ProductHistory.objects.filter(created_at__gte=start_of_month).order_by('-created_at')
            summary = defaultdict(lambda: {'count': 0, 'received_total': 0, 'price_total': 0})

            for i, product in enumerate(products, start=1):
                ws.append([
                    i,
                    product.name or '',
                    product.sku or '',
                    str(product.category) or '',
                    product.price_received or 0,
                    product.price or 0,
                    product.status or '',
                    product.count or 0,
                    localtime(product.created_at).strftime("%Y-%m-%d %H:%M:%S")
                ])

                status = product.status or 'no-status'
                count = product.count or 0
                received = product.price_received or 0
                price = product.price or 0

                summary[status]['count'] += count
                summary[status]['received_total'] += abs(count * received)
                summary[status]['price_total'] += abs(count * price)

            ws_sum = wb.create_sheet("Hisobot")
            ws_sum.append(["Amal turi", "Umumiy soni", "Olingan summa", "Sotuv summasi", "Soft foyda"])
            for cell in ws_sum["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            column_widths = [12, 20, 20, 20, 15]
            for i, width in enumerate(column_widths, start=1):
                col_letter = ws_sum.cell(row=1, column=i).column_letter
                ws_sum.column_dimensions[col_letter].width = width

            for status, data in summary.items():
                profit = data['price_total'] - data['received_total']
                ws_sum.append([
                    status,
                    data['count'],
                    data['received_total'],
                    data['price_total'],
                    profit
                ])

            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            filename = f"shu_oy_mahsulotlar_{today.strftime('%Y_%m_%d_%H_%M')}.xlsx"
            bot.send_document(
                user_id,
                excel_file,
                visible_file_name=filename,
                caption="📊 Shu oygi mahsulotlar va umumiy hisobot"
            )
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
            buttons = [
                types.InlineKeyboardButton("🔥 Bosh menyu", web_app=web_app_info),
                types.InlineKeyboardButton("📆 Shu oygi buyurtmalar", callback_data='orders_this_month'),
                types.InlineKeyboardButton("📋 Barcha buyurtmalar", callback_data='orders_all'),
                types.InlineKeyboardButton("📆 Shu oygi mahsulotlar", callback_data='products_this_month'),
                types.InlineKeyboardButton("📦 Barcha mahsulotlar", callback_data='products_all'),
            ]
            keyboard.add(*buttons)
            bot.send_message(user_id, "✅ Hisobot yuborildi. Menyudan birini tanlang:", reply_markup=keyboard)

        except Exception as e:
            print("Xatolik:", e)
            bot.send_message(user_id, "❌ Hisobotni yaratishda xatolik yuz berdi.")

    elif call.data == 'products_all':
        user_id = call.from_user.id
        try:
            wb = Workbook()
            del wb['Sheet']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            all_products = ProductHistory.objects.all().order_by('created_at')
            products_by_month = defaultdict(list)
            summary_total = defaultdict(lambda: {'count': 0, 'received_total': 0, 'price_total': 0})
            for product in all_products:
                month_key = localtime(product.created_at).strftime('%Y-%m')  # Masalan: '2025-06'
                products_by_month[month_key].append(product)
            for month, products in products_by_month.items():
                ws = wb.create_sheet(title=month)
                headers = ['#', 'Maxsulot nomi', 'SKU', "Bo'lim", 'Olingan narx', 'Sotuv narx',
                           'Amal', 'Soni', "Yaratilgan vaqt"]
                ws.append(headers)
                for cell in ws["1:1"]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                column_widths = [5, 20, 20, 20, 15, 15, 15, 15, 25]
                for i, width in enumerate(column_widths, start=1):
                    col_letter = ws.cell(row=1, column=i).column_letter
                    ws.column_dimensions[col_letter].width = width
                summary = defaultdict(lambda: {'count': 0, 'received_total': 0, 'price_total': 0})
                for i, product in enumerate(products, start=1):
                    count = product.count or 0
                    received = product.price_received or 0
                    price = product.price or 0
                    status = product.status or 'no-status'
                    ws.append([
                        i,
                        product.name or '',
                        product.sku or '',
                        str(product.category) or '',
                        received,
                        price,
                        status,
                        count,
                        localtime(product.created_at).strftime("%Y-%m-%d %H:%M:%S")
                    ])
                    summary[status]['count'] += count
                    summary[status]['received_total'] += abs(count * received)
                    summary[status]['price_total'] += abs(count * price)

                    summary_total[status]['count'] += count
                    summary_total[status]['received_total'] += abs(count * received)
                    summary_total[status]['price_total'] += abs(count * price)
                ws_sum = wb.create_sheet(title=f"{month} Hisobot")
                ws_sum.append(["Amal turi", "Umumiy soni", "Olingan summa", "Sotuv summasi", "Soft foyda"])
                for cell in ws_sum["1:1"]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                column_widths = [12, 20, 20, 20, 20]
                for i, width in enumerate(column_widths, start=1):
                    col_letter = ws_sum.cell(row=1, column=i).column_letter
                    ws_sum.column_dimensions[col_letter].width = width
                for status, data in summary.items():
                    profit = data['price_total'] - data['received_total']
                    ws_sum.append([status, data['count'], data['received_total'], data['price_total'], profit])
            ws_total = wb.create_sheet(title="Umumiy Hisobot")
            ws_total.append(["Amal turi", "Umumiy soni", "Olingan summa", "Sotuv summasi", "Soft foyda"])
            for cell in ws_total["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            column_widths = [12, 20, 20, 20, 20]
            for i, width in enumerate(column_widths, start=1):
                col_letter = ws_total.cell(row=1, column=i).column_letter
                ws_total.column_dimensions[col_letter].width = width
            grand_total = {'count': 0, 'received_total': 0, 'price_total': 0}
            total_profit = 0

            for status, data in summary_total.items():
                profit = data['price_total'] - data['received_total']
                total_profit += profit
                grand_total['count'] += data['count']
                grand_total['received_total'] += data['received_total']
                grand_total['price_total'] += data['price_total']
                ws_total.append([status, data['count'], data['received_total'], data['price_total'], profit])
            ws_total.append(["JAMI", grand_total['count'], grand_total['received_total'],
                             grand_total['price_total'], total_profit])
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            today = now()
            bot.send_document(
                user_id,
                excel_file,
                visible_file_name=f"barcha_maxsulotlar_oyma_oy_{today.strftime('%Y_%m_%d_%H_%M')}.xlsx",
                caption="📦 Barcha mahsulotlar (oyma-oy + umumiy hisobot)"
            )
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
            buttons = [
                types.InlineKeyboardButton("🔥 Bosh menyu", web_app=web_app_info),
                types.InlineKeyboardButton("📆 Shu oygi buyurtmalar", callback_data='orders_this_month'),
                types.InlineKeyboardButton("📋 Barcha buyurtmalar", callback_data='orders_all'),
                types.InlineKeyboardButton("📆 Shu oygi mahsulotlar", callback_data='products_this_month'),
                types.InlineKeyboardButton("📦 Barcha mahsulotlar", callback_data='products_all'),
            ]
            keyboard.add(*buttons)
            bot.send_message(user_id, "✅ Hisobot yuborildi. Menyudan birini tanlang:", reply_markup=keyboard)
        except Exception as e:
            print("Xatolik:", e)
            bot.send_message(user_id, "❌ Barcha mahsulotlar hisobotida xatolik yuz berdi.")


def send_telegram_message(message: str):
    try:
        users = User.objects.filter(is_superuser=True)
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
        buttons = [
            types.InlineKeyboardButton("🔥 Bosh menyu", web_app=web_app_info),
            types.InlineKeyboardButton("📆 Shu oygi buyurtmalar", callback_data='orders_this_month'),
            types.InlineKeyboardButton("📋 Barcha buyurtmalar", callback_data='orders_all'),
            types.InlineKeyboardButton("📆 Shu oygi mahsulotlar", callback_data='products_this_month'),
            types.InlineKeyboardButton("📦 Barcha mahsulotlar", callback_data='products_all'),
        ]
        keyboard.add(*buttons)
        for user in users:
            try:
                bot.send_message(user.user_telegram_id, message, parse_mode="HTML", reply_markup=keyboard)
            except exceptions:
                pass
    except Exception as e:
        print(f"Telegramga habar yuborishda xatolik: {e}")


def get_user_info(message):
    user_id = message.from_user.id
    user_data[user_id] = {'name': message.text}

    bot.send_message(user_id, "Telefon raqamingizni yuboring (+998991234567 shunday ko'rinishda):")
    bot.register_next_step_handler(message, get_user_phone)

def get_user_phone(message):
    user_id = message.from_user.id

    data = user_data.get(user_id)
    if len(message.text) == 13 and message.text[:4] == "+998":
        if data:
            user, created = User.objects.update_or_create(
                user_telegram_id=user_id,
                defaults={
                    'first_name': data['name'],
                    'last_name': message.from_user.first_name,
                    'telegram_username': message.from_user.username,
                    'phone_number': message.text,
                    'username': user_id
                }
            )
            bot.send_message(message.chat.id, "✅ Ro'yxatdan muvaffaqiyatli o'tdingiz!")
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            web_app_info = types.WebAppInfo(url="https://jinxinguz.netlify.app/#/")
            buttons = [
                types.InlineKeyboardButton("🔥 Mahsulotlar", web_app=web_app_info),
                types.InlineKeyboardButton("👤 Siz haqingizda", callback_data='user'),
                types.InlineKeyboardButton(" 📞 Bog'lanish", callback_data='support'),
            ]
            keyboard.add(*buttons)
            bot.send_message(user_id, "Assalomu alaykum! Tanlang:", reply_markup=keyboard)
            user_data.pop(user_id)
        else:
            bot.send_message(message.chat.id, "⚠️ Avval /start buyrug'ini yuboring.")
    else:
        bot.send_message(user_id, "Telefon raqamingizni xato yubordingiz qaytadan yuboring⁉️")
        bot.register_next_step_handler(message, get_user_phone)


class Command(BaseCommand):
    help = 'Telegram botni ishga tushuradi'

    def handle(self, *args, **kwargs):
        print("Bot ishga tushdi...")
        bot.infinity_polling()